from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook

# Set up Selenium WebDriver with headless Chrome
chrome_options = Options()
chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)

# Automatically manage ChromeDriver with webdriver_manager
service = Service(ChromeDriverManager().install())

driver = webdriver.Chrome(service=service, options=chrome_options)

# URL of the page you want to scrape
url = "https://www.alojamiento.io/property/apartamentos-centro-coló3n/BC-189483"
driver.get(url)

# Execute JavaScript to get the 'ScriptData' object
script_data = driver.execute_script("""
    return ScriptData;
""")

# Extract the required information from 'script_data'
site_url = script_data['config']['SiteUrl']
campaign_id = script_data['pageData']['CampaignId']
site_name = script_data['config']['SiteName']
browser = script_data['userInfo']['Browser']
country_code = script_data['userInfo']['CountryCode']
ip = script_data['userInfo']['IP']

# Load the existing Excel workbook (test_report.xlsx)
file_path = "test_report.xlsx"
wb = load_workbook(file_path)

# Create a new sheet for the scraped data
ws = wb.create_sheet(title="Scraped Data")

# Write headers to the new sheet
headers = ['SiteURL', 'CampaignID', 'SiteName', 'Browser', 'CountryCode', 'IP']
ws.append(headers)

# Write the extracted data to the new sheet
ws.append([site_url, campaign_id, site_name, browser, country_code, ip])

# Save the updated workbook
wb.save(file_path)

# Close the Selenium WebDriver
driver.quit()
