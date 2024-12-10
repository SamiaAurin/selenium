import openpyxl
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# Initialize the Excel workbook and sheets
def initialize_excel_report():
    workbook = openpyxl.Workbook()
    
    # Create separate sheets for each test
    sheet_h1 = workbook.active
    sheet_h1.title = "H1 Tag Test"
    sheet_h1.append(["URL", "Status", "Comments"])
    
    sheet_html_sequence = workbook.create_sheet(title="HTML Tag Sequence")
    sheet_html_sequence.append(["URL", "Status", "Comments"])
    
    sheet_image_alt = workbook.create_sheet(title="Image Alt Test")
    sheet_image_alt.append(["URL", "Status", "Comments"])
    
    # Make headers bold
    for sheet in [sheet_h1, sheet_html_sequence, sheet_image_alt]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
    
    return workbook, sheet_h1, sheet_html_sequence, sheet_image_alt

# H1 Tag Existence Test
def test_h1_tag(driver, url):
    driver.get(url)
    try:
        driver.find_element(By.TAG_NAME, "h1")
        return ("Pass", "H1 tag found")
    except:
        return ("Fail", "H1 tag missing")

# HTML Tag Sequence Test
def test_html_tag_sequence(driver, url):
    driver.get(url)
    headings = driver.find_elements(By.XPATH, "//h1 | //h2 | //h3 | //h4 | //h5 | //h6")
    heading_levels = [int(h.tag_name[1]) for h in headings]  # Extract the number from h1, h2, etc.

    for i in range(1, len(heading_levels)):
        if heading_levels[i] > heading_levels[i - 1] + 1:
            return ("Fail", f"Sequence broken: {heading_levels}")
    
    if not heading_levels:
        return ("Fail", "No headings found on the page")
    
    return ("Pass", f"Sequence correct: {heading_levels}")

# Image Alt Attribute Test
def test_image_alt_attribute(driver, url):
    driver.get(url)
    images = driver.find_elements(By.TAG_NAME, "img")
    missing_alt = [img.get_attribute("src") for img in images if not img.get_attribute("alt")]
    
    if missing_alt:
        return ("Fail", f"Missing alt attribute for {len(missing_alt)} images")
    return ("Pass", "All images have alt attributes")

# Main function to perform tests and save results
def main():
    # List of URLs to test

    urls = [
    "https://www.alojamiento.io/",
    "https://www.alojamiento.io/",
    "https://www.alojamiento.io/",
    "https://www.facebook.com/StaysTravel",
    "https://x.com/StaysTravel",
    "https://www.instagram.com/staystravel",
    "https://www.alojamiento.io/place-to-stay",
    "https://www.alojamiento.io/near-me?all=true",
    "https://www.alojamiento.io/addalisting",
    "https://www.alojamiento.io/about-us",
    "https://www.alojamiento.io/faq",
    "https://www.alojamiento.io/all/spain/valencian-community/alicante",
    "https://www.alojamiento.io/all/spain/catalonia/barcelona",
    "https://www.alojamiento.io/all/spain/valencian-community/benidorm",
    "https://www.alojamiento.io/all/spain/basque-country/bilbao",
    "https://www.alojamiento.io/all/spain/andalusia/cordoba",
    "https://www.alojamiento.io/all/spain/canary-islands/costa-adeje",
    "https://www.alojamiento.io/all/spain/andalusia/costa-del-sol",
    "https://www.alojamiento.io/all/spain/andalusia/granada",
    "https://www.alojamiento.io/all/spain/community-of-madrid/madrid/ibiza",
    "https://www.alojamiento.io/all/spain/community-of-madrid/madrid",
    "https://www.alojamiento.io/all/spain/andalusia/malaga",
    "https://www.alojamiento.io/all/spain/balearic-islands/mallorca-island",
    "https://www.alojamiento.io/all/spain/andalusia/marbella",
    "https://www.alojamiento.io/all/spain/canary-islands/san-bartolome-de-tirajana/maspalomas",
    "https://www.alojamiento.io/all/spain/catalonia/salou",
    "https://www.alojamiento.io/all/spain/san-sebastian",
    "https://www.alojamiento.io/all/spain/andalusia/seville",
    "https://www.alojamiento.io/all/spain/castilla-la-mancha/toledo",
    "https://www.alojamiento.io/all/spain/andalusia/torremolinos",
    "https://www.alojamiento.io/all/spain/valencian-community/valencia",
    "https://www.alojamiento.io/all/spain/aragon/zaragoza",
    "https://www.alojamiento.io/all/spain/community-of-madrid/madrid",
    "https://www.alojamiento.io/all/spain/catalonia/barcelona",
    "https://www.alojamiento.io/all/spain/andalusia/seville",
    "https://www.alojamiento.io/all/spain/valencian-community/valencia",
    "https://www.alojamiento.io/all/spain/basque-country/bilbao",
    "https://www.alojamiento.io/all/spain/valencian-community/alicante",
    "https://www.alojamiento.io/all/spain/balearic-islands/palma-de-mallorca",
    "https://www.alojamiento.io/all/spain/balearic-islands/menorca",
    "https://www.alojamiento.io/all/hotels",
    "https://www.alojamiento.io/all/andorra/andorra-la-vella/andorra-la-vella",
    "https://www.alojamiento.io/all/portugal/algarve",
    "https://www.alojamiento.io/all/portugal/porto-district/porto",
    "https://www.alojamiento.io/all/portugal/lisbon-district/lisbon",
    "https://www.alojamiento.io/all/morocco/marrakech-safi/marrakech",
    "https://www.alojamiento.io/all/spain/catalonia/barcelona",
    "https://www.alojamiento.io/all/spain/basque-country/bilbao",
    "https://www.alojamiento.io/all/turkey/istanbul",
    "https://www.alojamiento.io/all/spain/community-of-madrid/madrid",
    "https://www.alojamiento.io/all/spain/andalusia/seville",
    "https://www.alojamiento.io/all/italy/veneto/venice",
    "https://www.alojamiento.io/all/france/provence",
    "https://www.alojamiento.io/all/spain/balearic-islands/palma-de-mallorca",
    "https://www.alojamiento.io/all/spain/valencian-community/benidorm",
    "https://www.alojamiento.io/place-to-stay",
    "https://www.alojamiento.io/all/france/paris",
    "https://www.alojamiento.io/all/portugal",
    "https://www.alojamiento.io/all/italy",
    "https://www.alojamiento.io/all/mexico",
    "https://www.alojamiento.io/all/switzerland",
    "https://www.alojamiento.io/all/greece",
    "https://www.alojamiento.io/all/resorts",
    "https://www.alojamiento.io/",
    "https://www.alojamiento.io/addalisting",
    "https://www.alojamiento.io/about-us",
    "https://www.alojamiento.io/faq",
    "https://www.alojamiento.io/all/spain/cantabria",
    "https://www.alojamiento.io/all/spain/community-of-madrid/madrid",
    "https://www.alojamiento.io/all/spain/catalonia",
    "https://www.alojamiento.io/all/spain/galicia",
    "https://www.alojamiento.io/all/spain/andalusia/malaga",
    "https://www.alojamiento.io/all/spain/navarre",
    "https://www.alojamiento.io/all/spain/canary-islands/gran-canaria",
    "https://www.alojamiento.io/all/spain/canary-islands/tenerife",
    "https://www.alojamiento.io/all/usa/wyoming/basin",
    "https://www.alojamiento.io/all/spain/castile-and-leon",
    "https://www.alojamiento.io/all/spain/murcia",
    "https://www.alojamiento.io/all/spain/valencian-community/valencia-province",
    "https://www.alojamiento.io/all/spain/andalusia",
    "https://www.alojamiento.io/all/spain/canary-islands",
    "https://www.alojamiento.io/all/spain/balearic-islands",
    "https://www.alojamiento.io/all/switzerland",
    "https://www.alojamiento.io/all/spain/community-of-madrid/madrid",
    "https://www.alojamiento.io/all/spain/catalonia/costa-brava",
    "https://www.alojamiento.io/all/italy",
    "https://www.alojamiento.io/all/france",
    "https://www.alojamiento.io/all/mexico",
    "https://www.alojamiento.io/all/morocco",
    "https://www.alojamiento.io/all/andorra",
    "https://www.alojamiento.io/all/greece",
    "https://www.alojamiento.io/all/mexico/quintana-roo/cancun",
    "https://www.alojamiento.io/all/dominican-republic/la-altagracia/punta-cana",
    "https://www.alojamiento.io/all/mexico/quintana-roo/playa-del-carmen",
    "https://www.alojamiento.io/all/puerto-rico/san-juan",
    "https://www.alojamiento.io/all/cuba/province-of-havana",
    "https://www.alojamiento.io/all/mexico/jalisco/puerto-vallarta",
    "https://www.alojamiento.io/all/mexico/baja-california-sur/cabo-san-lucas",
    "https://www.alojamiento.io/all/brazil/southeast-region/rio-de-janeiro",
    "https://www.alojamiento.io/all/argentina/buenos-aires",
    "https://www.alojamiento.io/all/mexico/quintana-roo/tulum",
    "https://www.alojamiento.io/all/colombia/medellin",
    "https://www.alojamiento.io/all/colombia/bolivar/cartagena",
    "https://www.alojamiento.io/all/peru/lima",
    "https://www.alojamiento.io/all/chile/santiago-metropolitan/santiago",
    "https://www.alojamiento.io/all/costa-rica/san-jose",
    "https://www.alojamiento.io/all/uruguay/montevideo",
    "https://www.alojamiento.io/all/panama/panama/panama-city",
    "https://www.alojamiento.io/all/brazil/south-region/florianopolis",
    "https://www.alojamiento.io/all/ecuador/pichincha/quito",
    "https://www.alojamiento.io/all/guatemala/sacatepequez/antigua-guatemala",
    "https://www.alojamiento.io/all/brazil/northeast-region/salvador",
    "https://www.alojamiento.io/all/peru/cusco/cusco",
    "https://www.alojamiento.io/all/honduras/bay-islands/roatan",
    "https://www.alojamiento.io/all/bolivia/la-paz",
    "https://www.alojamiento.io/all/ecuador/galapagos",
    "https://www.alojamiento.io/all/mexico/guanajuato/san-miguel-de-allende",
    "https://www.alojamiento.io/all/argentina/mendoza",
    "https://www.alojamiento.io/all/colombia/magdalena/santa-marta",
    "https://www.alojamiento.io/all/belize/belize-district/belize-city",
    "https://www.alojamiento.io/all/argentina/rio-negro/san-carlos-de-bariloche",
    "https://www.facebook.com/StaysTravel",
    "https://x.com/StaysTravel",
    "https://www.instagram.com/staystravel",
    "https://www.alojamiento.io/privacy-policy",
    "https://www.alojamiento.io/site-terms",
    "https://www.alojamiento.io/site-map",
    "https://www.alojamiento.io/",
    "https://www.travelai.com/",
    "https://www.alojamiento.io/privacy-policy#site-cookie-policy",
    "https://www.petfriendly.io/",
    "https://www.onedegreeleft.com/",
    "https://www.alojamiento.io/site-terms"
 
    ]
    
    # Set up the WebDriver
    options = Options()
    #options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    
    # Initialize the Excel report
    workbook, sheet_h1, sheet_html_sequence, sheet_image_alt = initialize_excel_report()
    
    # Loop through each URL and run all tests
    for url in urls:
        print(f"Testing URL: {url}")
        
        # H1 Tag Test
        status, comments = test_h1_tag(driver, url)
        sheet_h1.append([url, status, comments])
        
        # HTML Tag Sequence Test
        status, comments = test_html_tag_sequence(driver, url)
        sheet_html_sequence.append([url, status, comments])
        
        # Image Alt Attribute Test
        status, comments = test_image_alt_attribute(driver, url)
        sheet_image_alt.append([url, status, comments])
    
    # Save the Excel report
    workbook.save("test_report_allURL.xlsx")
    print("All tests completed. Report saved as 'test_report_allURL.xlsx'.")
    
    # Quit the WebDriver
    driver.quit()

# Execute the script
if __name__ == "__main__":
    main()
