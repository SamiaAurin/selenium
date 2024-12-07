from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
import os
import time


def save_results_to_excel(initial_prices, updated_prices, test_results, initial_availability_price, updated_availability_price, availability_result):
    # Define the Excel file name
    excel_file = "test_report.xlsx"

    # Check if the file exists
    if os.path.exists(excel_file):
        # Load the existing workbook
        workbook = load_workbook(excel_file)
        # Create a new sheet
        sheet = workbook.create_sheet(title="Currency Change Test")
    else:
        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Currency Change Test"

    # Add header row
    header = ["Card Number", "Initial Price", "Updated Price", "Test Result"]
    sheet.append(header)

    # Apply bold styling to the header row
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Write the card results
    for i in range(len(initial_prices)):
        sheet.append([f"Card {i + 1}", initial_prices[i], updated_prices[i], test_results[i]])

    # Add the availability price test results to the last row
    sheet.append([
        "Availability Price Test",
        initial_availability_price,
        updated_availability_price,
        availability_result
    ])

    # Save the workbook
    workbook.save(excel_file)
    print(f"\nTest results saved to '{excel_file}'.")


def test_currency_change_for_cards(driver, url):
    # Open the webpage
    driver.get(url)

    # Capture the initial value of the availability price
    availability_price_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'js-default-price'))
    )
    initial_availability_price = availability_price_element.text.strip()
    print(f"Initial Availability Price: {initial_availability_price}")

    # Wait for the price elements to load in cards
    print("Waiting for price elements...")
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, 'js-price-value'))
    )
    price_elements = driver.find_elements(By.CLASS_NAME, 'js-price-value')
    initial_prices = [elem.text for elem in price_elements]
    print(f"Found {len(initial_prices)} price elements.")

    # Print initial prices
    print("Initial Prices:")
    for i, price in enumerate(initial_prices):
        print(f"Card {i + 1}: {price}")

    # Scroll to the footer section to locate the currency dropdown
    print("Waiting for currency dropdown to become present...")
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, 'js-currency-sort-footer'))
    )
    footer_currency_element = driver.find_element(By.ID, 'js-currency-sort-footer')
    driver.execute_script("arguments[0].scrollIntoView(true);", footer_currency_element)
    time.sleep(1)  # Wait for the scroll to complete

    # Click on the currency dropdown
    try:
        print("Waiting for currency dropdown to become clickable...")
        currency_dropdown = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.ID, 'js-currency-sort-footer'))
        )
        currency_dropdown.click()
        print("Currency dropdown clicked.")

        # Wait for the dropdown to become visible
        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located(
                (By.XPATH, "//div[@class='footer-section']//div[@class='footer-currency-dd']//ul[@class='select-ul']//li[.//div[@class='option']//p[contains(text(), '$ (USD)')]]")
            )
        )
        print("Dropdown menu is visible.")

        # Find the USD option using XPath
        usd_option = driver.find_element(
            By.XPATH, "//div[@class='footer-section']//div[@class='footer-currency-dd']//ul[@class='select-ul']//li[.//div[@class='option']//p[contains(text(), '$ (USD)')]]"
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", usd_option)
        print("USD option found and scrolled into view.")
        try:
            usd_option.click()
        except Exception:
            print("Click intercepted, using JavaScript to click USD option.")
            driver.execute_script("arguments[0].click();", usd_option)

        # Wait for the availability price to update
        WebDriverWait(driver, 50).until(
            EC.text_to_be_present_in_element((By.ID, 'js-default-price'), '$')
        )
        updated_availability_price = availability_price_element.text.strip()
        print(f"Updated Availability Price: {updated_availability_price}")

        # Wait for the prices to update in cards
        WebDriverWait(driver, 50).until(
            EC.text_to_be_present_in_element((By.CLASS_NAME, 'js-price-value'), '$')
        )
        print("Prices updated with USD.")

    except Exception as e:
        print(f"Error interacting with currency dropdown: {e}")
        return

    # Capture updated prices after currency change
    try:
        updated_prices = [elem.text for elem in driver.find_elements(By.CLASS_NAME, 'js-price-value')]

        print("Updated Prices (After Currency Change to USD):")
        for i, price in enumerate(updated_prices):
            print(f"Card {i + 1}: {price}")

        test_results = []
        for i in range(len(initial_prices)):
            if initial_prices[i] != updated_prices[i]:
                test_results.append(f"PASS (Currency changed successfully)")
            else:
                test_results.append(f"FAIL (Currency did not change)")

        # Verify if availability price has changed
        if "$" in updated_availability_price:
            availability_result = "PASS (Currency changed successfully)"
        else:
            availability_result = "FAIL (Currency did not change)"

        print(f"\nAvailability Price Test Result: {availability_result}")

        # Save results to Excel
        save_results_to_excel(
            initial_prices,
            updated_prices,
            test_results,
            initial_availability_price,
            updated_availability_price,
            availability_result
        )

    except Exception as e:
        print(f"Error capturing updated prices: {e}")

# Main execution
if __name__ == "__main__":
    # Set Chrome options
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')  # Uncomment to run in headless mode
    options.add_argument('--disable-gpu')

    # Use Service to pass the executable path
    service = Service(ChromeDriverManager().install())

    # Initialize the WebDriver
    driver = webdriver.Chrome(service=service, options=options)

    try:
        test_currency_change_for_cards(driver, "https://www.alojamiento.io/property/apartamentos-centro-coló3n/BC-189483")
    finally:
        driver.quit()
