import openpyxl
from openpyxl.styles import Font
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import time

# Initialize the Excel report
def initialize_excel_report():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test Report"
    sheet.append(["Page URL", "Test Name", "Status", "Comments"])  
    for cell in sheet[1]:
        cell.font = Font(bold=True)  
    return workbook, sheet

# H1 Tag Existence Test
def test_h1_tag(driver, url):
    driver.get(url)
    try:
        h1_tag = driver.find_element(By.TAG_NAME, "h1")
        return ("H1 Tag Existence", "Pass", "H1 tag found")
    except:
        return ("H1 Tag Existence", "Fail", "H1 tag missing")

# HTML Tag Sequence Test
def test_html_tag_sequence(driver, url):
    driver.get(url)
    headings = driver.find_elements(By.XPATH, "//h1 | //h2 | //h3 | //h4 | //h5 | //h6")
    heading_levels = [int(h.tag_name[1]) for h in headings]  # Extract the number from h1, h2, etc.

    # Check if the sequence is correct
    for i in range(1, len(heading_levels)):
        if heading_levels[i] > heading_levels[i - 1] + 1:  # If a level is skipped
            return ("HTML Tag Sequence", "Fail", f"Sequence broken: {heading_levels}")
    
    if not heading_levels:
        return ("HTML Tag Sequence", "Fail", "No headings found on the page")
    
    return ("HTML Tag Sequence", "Pass", f"Sequence correct: {heading_levels}")


# Image Alt Attribute Test
def test_image_alt_attribute(driver, url):
    driver.get(url)
    images = driver.find_elements(By.TAG_NAME, "img")  # Find all <img> tags on the page
    missing_alt = []  # List to store images missing the alt attribute

    for img in images:
        alt_text = img.get_attribute("alt")
        if not alt_text:  # Check if alt is missing or empty
            missing_alt.append(img.get_attribute("src"))  # Store the image source for reference

    if missing_alt:
        return ("Image Alt Attribute", "Fail", f"Missing alt attribute for {len(missing_alt)} images")
    else:
        return ("Image Alt Attribute", "Pass", "All images have alt attributes")


# URL Status Code Test
def test_url_status_code(driver, url, sheet_urls):
    driver.get(url)
    # Find only <a> tags within the body of the page
    body = driver.find_element(By.TAG_NAME, 'body')
    links = body.find_elements(By.TAG_NAME, "a")  
    broken_links = []  

    for link in links:
        href = link.get_attribute("href")  
        if href:  # Ensure the link is not empty
            try:
                response = requests.head(href, allow_redirects=True)  # Check the URL status
                if response.status_code == 404:  # Log only broken links
                    broken_links.append(href)
                    sheet_urls.append([href, "Fail", "404 Not Found"])
            except requests.RequestException as e:  # Log request errors
                broken_links.append(href)
                sheet_urls.append([href, "Fail", f"Request error: {str(e)}"])
        else:
            sheet_urls.append(["Empty Link", "Fail", "No href attribute found"])  # Log missing href attributes

    # Return the test result
    if broken_links:
        return ("URL Status Code", "Fail", f"{len(broken_links)} broken links found")
    else:
        return ("URL Status Code", "Pass", "No broken links found")


# Currency filtering and ensure property tiles currency changed
def test_currency_change_for_all(workbook, driver, url):
    
    sheet = workbook.create_sheet(title="Currency Change Results")
    sheet.append([
        "Currency", "Card", "Initial Price", "Updated Price", "Card Test Result",
        "Initial Availability Price", "Updated Availability Price", "Availability Test Result"
    ])  
    for cell in sheet[1]:
        cell.font = Font(bold=True)  

    # Open the webpage
    driver.get(url)

    # Capture the initial value of the availability price
    availability_price_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'js-default-price'))
    )
    initial_availability_price = availability_price_element.text.strip()
    print(f"Initial Availability Price: {initial_availability_price}")

    # Wait for the price elements to load in cards
    print("Waiting for price elements...")
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, 'js-price-value'))
    )
    price_elements = driver.find_elements(By.CLASS_NAME, 'js-price-value')
    initial_prices = [elem.text for elem in price_elements]
    print(f"Found {len(initial_prices)} price elements.")

    # Scroll to the footer section to locate the currency dropdown
    print("Waiting for currency dropdown to become present...")
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.ID, 'js-currency-sort-footer'))
    )
    footer_currency_element = driver.find_element(By.ID, 'js-currency-sort-footer')
    driver.execute_script("arguments[0].scrollIntoView(true);", footer_currency_element)
    time.sleep(1)  # Wait for the scroll to complete

    # Click on the currency dropdown
    currency_dropdown = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.ID, 'js-currency-sort-footer'))
    )
    currency_dropdown.click()
    print("Currency dropdown clicked.")

    # Fetch all available currency options
    currency_options = driver.find_elements(By.XPATH, "//div[@class='footer-section']//div[@class='footer-currency-dd']//ul[@class='select-ul']//li")
    print(f"Found {len(currency_options)} currency options.")

    # Iterate over each currency option
    for currency_option in currency_options:
        currency_text = currency_option.text.strip()
        
        # Skip empty or invalid options
        if not currency_text:
            print("Skipping an empty currency option.")
            continue

        print(f"\nTesting currency: {currency_text}")

        # Scroll to and click on the currency option
        driver.execute_script("arguments[0].scrollIntoView(true);", currency_option)
        time.sleep(1)
        try:
            currency_option.click()
        except Exception:
            driver.execute_script("arguments[0].click();", currency_option)
        
        # Wait for the availability price to update
        WebDriverWait(driver, 60).until(
            EC.text_to_be_present_in_element((By.ID, 'js-default-price'), currency_text.split()[0])
        )
        # Capture the update value of the availability price
        availability_price_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'js-default-price'))
        )
        updated_availability_price = availability_price_element.text.strip()
        print(f"Updated Availability Price: {updated_availability_price}")

        # Wait for the prices to update in the cards
        WebDriverWait(driver, 60).until(
            EC.text_to_be_present_in_element((By.CLASS_NAME, 'js-price-value'), currency_text.split()[0])
        )
        updated_price_elements = driver.find_elements(By.CLASS_NAME, 'js-price-value')
        updated_prices = [elem.text for elem in updated_price_elements]

        # Compare initial and updated prices
        for i in range(len(initial_prices)):

            card_result = "PASS (Currency changed successfully)" if currency_text.split()[0] in updated_prices[i] else "FAIL (Currency did not change)"
            availability_result = "PASS (Currency changed successfully)" if currency_text.split()[0] in updated_availability_price else "FAIL (Currency did not change)"
            
            # Write the results into the sheet
            if i == 0:  # Write availability prices only for the first card of the currency
                sheet.append([
                    currency_text, 
                    f"Card {i + 1}",
                    initial_prices[i], 
                    updated_prices[i], 
                    card_result,
                    initial_availability_price,
                    updated_availability_price,
                    availability_result
                ])
            else:
                sheet.append([
                    currency_text, 
                    f"Card {i + 1}",
                    initial_prices[i], 
                    updated_prices[i], 
                    card_result,
                    "",  # No need to repeat availability prices
                    "",
                    ""
                ])

        # Reopen the dropdown for the next currency
        currency_dropdown = driver.find_element(By.ID, 'js-currency-sort-footer')
        driver.execute_script("arguments[0].scrollIntoView(true);", currency_dropdown)
        currency_dropdown.click()
        currency_options = driver.find_elements(By.XPATH, "//div[@class='footer-section']//div[@class='footer-currency-dd']//ul[@class='select-ul']//li")

    return "Currency Change Test", "Pass", "Currency test results saved in Excel"


# Write the Scraped Data (Sheet 4)
def write_scraped_data(workbook, driver, url):
    sheet = workbook.create_sheet(title="Scraped Data")
    sheet.append(["Site URL", "Campaign ID", "Site Name", "Browser", "Country Code", "IP"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)  

    # Navigate to the URL
    driver.get(url)
    try:
        # Execute JavaScript to get the 'ScriptData' object
        script_data = driver.execute_script("""
            return ScriptData;
        """)

        site_url = script_data['config']['SiteUrl']
        campaign_id = script_data['pageData']['CampaignId']
        site_name = script_data['config']['SiteName']
        browser = script_data['userInfo']['Browser']
        country_code = script_data['userInfo']['CountryCode']
        ip = script_data['userInfo']['IP']

        sheet.append([site_url, campaign_id, site_name, browser, country_code, ip])    
    except Exception as e:
        sheet.append(["N/A", "N/A", "N/A", "N/A", "N/A", "N/A", f"Error: {str(e)}"])
        print(f"Error extracting ScriptData: {e}")

    return "Scraped Data", "Pass" if script_data else "Fail", "Data written to sheet" if script_data else "No data found"


# Main function to run the tests and generate the Excel report
def main():
    # Set up the WebDriver
    options = webdriver.ChromeOptions()
    #options.add_argument('--headless')  # Uncomment it if you want it to run in headless mode.
    options.add_argument('--disable-gpu')  

    # Use Service to pass the executable path
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Test site URL
    url = "https://www.alojamiento.io/property/apartamentos-centro-coló3n/BC-189483"
    
    # Initialize the Excel report
    workbook, sheet = initialize_excel_report()
    
    # Create a new sheet for URL status results
    sheet_urls = workbook.create_sheet(title="URL Status")
    sheet_urls.append(["URL", "Status", "Comments"])  # Header row
    for cell in sheet_urls[1]:
        cell.font = Font(bold=True)  # Make headers bold
    
    # Run the H1 tag test
    test_name, status, comments = test_h1_tag(driver, url)
    sheet.append([url, test_name, status, comments])
    
    # Run the HTML tag sequence test
    test_name, status, comments = test_html_tag_sequence(driver, url)
    sheet.append([url, test_name, status, comments])
    
    # Run the image alt attribute test
    test_name, status, comments = test_image_alt_attribute(driver, url)
    sheet.append([url, test_name, status, comments])
    
    # Run the URL status code test
    test_name, status, comments = test_url_status_code(driver, url, sheet_urls)
    sheet.append([url, test_name, status, comments])
    
    # Run the Currency Change test
    test_name, status, comments = test_currency_change_for_all(workbook, driver, url)
    sheet.append([url, test_name, status, comments])


    # Run the ScriptData test
    test_name, status, comments = write_scraped_data(workbook, driver, url)
    sheet.append([url, test_name, status, comments])

    # Save the report
    workbook.save("TestReports_All.xlsx")
    print("Test completed. Report saved as 'TestReports_All.xlsx'.")
    
    # Quit the WebDriver
    driver.quit()

# Execute the script
if __name__ == "__main__":
    main()